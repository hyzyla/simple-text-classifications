try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.cell import get_column_letter
except ImportError:
    print("Please install openpyxl using following command: \n"
          "pip install openpyxl\n"
          "(Best practice is to do this in virtual environment)\n")
import argparse
from gender_predictor import get_predictor_from_file
import csv


def predict_from_xlsx(args):
    gp = get_predictor_from_file('data/female.txt', 'data/male.txt')
    path = args['path']
    sheet = args['sheet']
    name_column = args['name_column']
    gender_column = args['gender_column']
    row_offset = args['offset']
    write_path = args['write_path']

    wb = load_workbook(filename=path)
    sheet_ranges = wb[sheet]

    if gender_column == 'last':
        last_column = get_column_letter(sheet_ranges.max_column + 1)
        gender_column = last_column

    genders = []
    for cell in sheet_ranges[name_column][row_offset:]:
        name = cell.value
        gender_predicted = gp.get_gender(name, use_predictor=True)
        genders.append(gender_predicted)

    for cell, gender in zip(sheet_ranges[gender_column][row_offset:], genders):
        cell.value = gender

    wb.save(path)
    wb.close()

def predict_from_csv(args):
    gp = get_predictor_from_file('data/female.txt', 'data/male.txt')
    path = args['path']
    row_offset = args['offset']
    write_path = args['write_path']
    write_path = write_path if write_path else path
    rows = []
    with open(path) as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if isinstance(row, list):
                row = row[0]
            rows.append([row, gp.get_gender(row)])

    with open(write_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(rows)



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Predict gender for given name')
    parser.add_argument('path', metavar='path', type=str,
                        help='path to xlsx file for reading')
    parser.add_argument('--sheet', type=str, metavar='sheet', default='Sheet1',
                        help='name of sheet')
    parser.add_argument('--write-path', type=str, metavar='wp',
                        help='path to xlsx file for writing. If empty - path of read file')
    parser.add_argument('--name-column', type=str, default='A', metavar='nc',
                        help='column with names')
    parser.add_argument('--gender-column', type=str, default='last', metavar='gc',
                        help='column for writing genders')
    parser.add_argument('--offset', type=int, default=1, metavar='o',
                        help='how rows will be skipped from the top of spreadsheet')
    args = parser.parse_args()
    args = vars(args)

    if args['path'][-4:] == 'xlsx':
        predict_from_xlsx(args)
    elif args['path'][-3:] == 'csv':
        predict_from_csv(args)
    else:
        print('Unknown format {}'.format(args['path']))
