import random
import re

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter

from nltk.stem.snowball import SnowballStemmer
from nltk.corpus import stopwords

import numpy as np

from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline
import sys

def async_print(*argv, **kwards):
    print(*argv, **kwards)
    sys.stdout.flush()

def print_tabular(data):
    header  = ['Product', 'Class']
    row_format ="{:>15}" * (len(header) + 1)
    async_print(row_format.format("", *header))
    for head, row in zip(header, data):
        async_print(row_format.format(head, *row))



class TextCleaner:
    
    def cleaner(self, text):
        
        stop_words = set(stopwords.words('finnish'))
        full_text = text.lower()

        words = [full_text]
        return ' '.join(words)
    
    def transform(self, X):
        # Convert to ndarray 
        X_ndarray = np.array(X)

        # Vectorise function for cleaning text, only one column with text support
        func = np.vectorize(self.cleaner)
        
        # Return result
        return func(X_ndarray)

        
    def fit(self, X, y=None):
        return self

def write_result(path, result, sheet='Sheet1', category_column='last', offset=1):
    wb = load_workbook(filename=path)
    sheet_ranges = wb[sheet]

    if category_column == 'last':
        last_column = get_column_letter(sheet_ranges.max_column + 1)
        category_column = last_column



    for cell, product_and_category in zip(sheet_ranges[category_column][offset:], result):
        cell.value = str(product_and_category[1])
        
    wb.save(path)
    wb.close()

def write_predicted(result):

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for items in result:
        ws.append([str(i) for i in items])

    wb.save('predicted.xlsx') 
    wb.close()

def read_examples(path, sheet='Sheet1', product_column='A', category_column='B', offset=1):
    wb = load_workbook(path)
    sheet = wb[sheet]
    product_names = sheet[product_column]
    categories = sheet[category_column]
    cells = [product_names, categories]
    
    data = [[cell.value for cell in row] for row in cells]

    # [[1,2], [1, 2] ...]
    data = list(zip(*data))

    data = np.array(data)
    data = data[offset:]
    
    return data

def read_notclassified(path, sheet='Sheet1', product_column='A', offset=1):
    wb = load_workbook(path)
    sheet = wb[sheet]
    
    product_names = sheet[product_column]
    data = [c.value for c in product_names]
    
    data = data[offset:]
    return np.array(data)

def correct_decoding(data):
    func = np.vectorize(lambda s: s.replace('Ã¤', 'ä').replace('Ã¶', 'ö').replace('â‚¬', '').strip())
    return func(data)

def get_model(examples):
    # Random shuffiling data
    examples = examples.copy()
    np.random.seed = 0
    np.random.shuffle(examples)

    # X - is names, y - classes for predict
    X, y = examples[:, 0], examples[:, 1]

    # getting all finnish stop words
    stop_words = stopwords.words('finnish')

    # Create pipline 
    pipline = Pipeline([
        ('cln', TextCleaner()),
        ('vect', CountVectorizer(stop_words=stop_words, max_df=1e-1)),
        ('tfidf', TfidfTransformer(use_idf=True, smooth_idf=True)),
        ('clf', MultinomialNB(alpha=1e-6, fit_prior=True)),
    ])

    # Train model
    pipline.fit(X, y) 
    async_print('\tNaive Bayes accuracy on train data: {}'.format(pipline.score(X, y)))
    return pipline

finnish_stem = SnowballStemmer('finnish', ignore_stopwords=True)
stop_words = set(stopwords.words('finnish'))

def cleaner(text):
        
        full_text = text.lower()
    
        # Delete all chars that is not letter
        full_text = re.sub('[^a-z^ä^ö^é]', ' ', full_text)
        
        # Skip all words smaller than 2 symbols
        words = [word for word in full_text.split() if len(word) > 1]

        # Detele stop words 
        words = [word for word in words if word not in stop_words and word not in ['min', 'lle']]

        # Steaming words
        words = [finnish_stem.stem(word) for word in words]
    
        return ' '.join(words)
    
def contains_words(text, joined_words):
    words = joined_words.split()
    return all(w in text for w in words)

def get_category(matched):
    values, counts  = np.unique(matched[:, 1], return_counts=True)
    ind = np.argmax(counts)
    return values[ind]

contains_words = np.vectorize(contains_words)
cleaner_vect = np.vectorize(cleaner)

if __name__ == "__main__":
    test = numpy.array('')